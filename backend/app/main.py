import base64
import hashlib
import hmac
import io
import json
import os
import secrets
import socket
import threading
import time
import urllib.error
import uuid
from datetime import datetime, timedelta, timezone
from types import SimpleNamespace

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from psycopg.errors import UniqueViolation

from app import legacy
from app.db import (
    create_session,
    create_user_with_workspace,
    create_workspace,
    get_cache_entry,
    get_latest_run_refinement,
    get_local_identity,
    get_run,
    get_session_user,
    initialize_database,
    list_runs,
    list_stages,
    list_workspaces_for_user,
    revoke_session,
    save_cache_entry,
    save_export_metadata,
    save_run_refinement,
    save_run,
    save_stage,
    touch_user_login,
    update_run,
    user_can_access_workspace,
)


SESSION_COOKIE = "bdcomps_session"
SESSION_DAYS = int(os.environ.get("SESSION_DAYS", "14"))
PROMPT_VERSION = os.environ.get("PROMPT_VERSION", "2026-06-10")
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-5.5")


app = FastAPI(title="BeOne Comps API", version="0.2.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:4174", "http://localhost:4174"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class PullRequest(BaseModel):
    prompt: str = ""
    scope: dict = Field(default_factory=dict)
    mode: str = "standard"
    use_cache: bool = True


class AuthRequest(BaseModel):
    email: str
    password: str = Field(min_length=8)
    display_name: str | None = None


class LoginRequest(BaseModel):
    email: str
    password: str


class WorkspaceRequest(BaseModel):
    name: str = Field(min_length=1, max_length=120)


class RefinementRequest(BaseModel):
    instruction: str = Field(min_length=1, max_length=4000)


class PersistedActivityLog(list):
    def __init__(self, run_id: int, initial_events: list[dict]):
        super().__init__(initial_events)
        self.run_id = run_id

    def append(self, event):
        super().append(event)
        update_run(self.run_id, "running", activity_log=list(self))


@app.on_event("startup")
def startup():
    initialize_database()


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    iterations = 260_000
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return "pbkdf2_sha256${}${}${}".format(
        iterations,
        base64.urlsafe_b64encode(salt).decode("ascii"),
        base64.urlsafe_b64encode(digest).decode("ascii"),
    )


def verify_password(password: str, encoded: str | None) -> bool:
    if not encoded:
        return False
    try:
        algorithm, iterations_text, salt_text, digest_text = encoded.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        salt = base64.urlsafe_b64decode(salt_text.encode("ascii"))
        expected = base64.urlsafe_b64decode(digest_text.encode("ascii"))
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(iterations_text))
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def hash_session_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def set_session_cookie(response: Response, token: str, expires_at: datetime):
    response.set_cookie(
        SESSION_COOKIE,
        token,
        httponly=True,
        samesite="lax",
        secure=os.environ.get("COOKIE_SECURE", "0") == "1",
        expires=expires_at,
        path="/",
    )


def clear_session_cookie(response: Response):
    response.delete_cookie(SESSION_COOKIE, path="/")


def get_current_user(request: Request):
    token = request.cookies.get(SESSION_COOKIE)
    if not token:
        raise HTTPException(status_code=401, detail="Authentication required.")
    user = get_session_user(hash_session_token(token))
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required.")
    return user


def first_workspace_for_user(user):
    workspaces = list_workspaces_for_user(user["id"])
    if not workspaces:
        return create_workspace(user["id"], f"{user['email'].split('@')[0]}'s Workspace")
    return workspaces[0]


def require_workspace(workspace_id: int, user):
    workspace = user_can_access_workspace(user["id"], workspace_id)
    if not workspace:
        raise HTTPException(status_code=404, detail="Workspace not found.")
    return workspace


def normalize_mode(mode: str | None):
    mode = (mode or "standard").strip().lower()
    return mode if mode in {"draft", "standard", "memo-ready"} else "standard"


def stable_cache_key(prompt: str, scope: dict, mode: str):
    payload = {"prompt": prompt or "", "scope": scope or {}, "mode": mode}
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def public_user(user):
    return {"id": user["id"], "email": user["email"], "display_name": user.get("display_name")}


def public_workspace(workspace):
    return {
        "id": workspace["id"],
        "name": workspace["name"],
        "type": workspace["type"],
        "role": workspace.get("role", "owner"),
        "owner_user_id": workspace.get("owner_user_id"),
    }


@app.get("/health")
def health():
    return {"ok": True, "service": "backend"}


@app.post("/api/auth/signup")
def signup(request: AuthRequest, response: Response):
    try:
        user, workspace = create_user_with_workspace(
            request.email,
            hash_password(request.password),
            request.display_name,
        )
    except UniqueViolation:
        raise HTTPException(status_code=409, detail="An account already exists for that email.")
    token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=SESSION_DAYS)
    create_session(user["id"], hash_session_token(token), expires_at)
    set_session_cookie(response, token, expires_at)
    return {"ok": True, "user": public_user(user), "workspaces": [public_workspace(workspace)], "active_workspace_id": workspace["id"]}


@app.post("/api/auth/login")
def login(request: LoginRequest, response: Response):
    identity = get_local_identity(request.email)
    if not identity or not verify_password(request.password, identity["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password.")
    token = secrets.token_urlsafe(32)
    expires_at = datetime.now(timezone.utc) + timedelta(days=SESSION_DAYS)
    create_session(identity["id"], hash_session_token(token), expires_at)
    touch_user_login(identity["id"])
    set_session_cookie(response, token, expires_at)
    workspaces = list_workspaces_for_user(identity["id"])
    return {
        "ok": True,
        "user": public_user(identity),
        "workspaces": [public_workspace(item) for item in workspaces],
        "active_workspace_id": workspaces[0]["id"] if workspaces else None,
    }


@app.post("/api/auth/logout")
def logout(request: Request, response: Response):
    token = request.cookies.get(SESSION_COOKIE)
    if token:
        revoke_session(hash_session_token(token))
    clear_session_cookie(response)
    return {"ok": True}


@app.get("/api/auth/me")
def me(user=Depends(get_current_user)):
    workspaces = list_workspaces_for_user(user["id"])
    return {
        "ok": True,
        "user": public_user(user),
        "workspaces": [public_workspace(item) for item in workspaces],
        "active_workspace_id": workspaces[0]["id"] if workspaces else None,
    }


@app.get("/api/workspaces")
def workspaces(user=Depends(get_current_user)):
    return {"ok": True, "workspaces": [public_workspace(item) for item in list_workspaces_for_user(user["id"])]}


@app.post("/api/workspaces")
def add_workspace(request: WorkspaceRequest, user=Depends(get_current_user)):
    return {"ok": True, "workspace": public_workspace(create_workspace(user["id"], request.name.strip()))}


@app.get("/api/workspaces/{workspace_id}")
def workspace_detail(workspace_id: int, user=Depends(get_current_user)):
    return {"ok": True, "workspace": public_workspace(require_workspace(workspace_id, user))}


def upload_to_legacy_field(file: UploadFile):
    file.file.seek(0)
    return SimpleNamespace(filename=file.filename, file=io.BytesIO(file.file.read()))


def prepare_scope(prompt: str, scope: dict, files: list[UploadFile] | None = None):
    log = []
    request_id = uuid.uuid4().hex[:10]
    legacy.activity(log, "Backend request started", "Handling Run pull request.", request_id=request_id)
    if files:
        datasets, errors = legacy.parse_uploaded_datasets([upload_to_legacy_field(file) for file in files], log)
        scope["uploadedDatasets"] = [
            {"name": item["name"], "type": item["type"], "row_count": item["row_count"]}
            for item in datasets
        ]
        scope["uploadedDatasetErrors"] = errors
        scope["uploadedDatasetRows"] = [
            {"dataset": item["name"], "type": item["type"], "row_count": item["row_count"], "rows": item["rows"]}
            for item in datasets
        ]
        legacy.activity(
            log,
            "Candidate input prepared",
            "Uploaded database exports will be treated as candidate leads requiring source verification.",
            datasets=len(datasets),
            dataset_errors=len(errors),
            parsed_rows=sum(item["row_count"] for item in datasets),
        )
    else:
        legacy.activity(log, "Candidate input prepared", "No uploaded datasets in this request.", datasets=0, parsed_rows=0)
    return request_id, prompt, scope, log


def complete_from_cache(run_id, result, log):
    legacy.activity(log, "Cache hit", "A completed equivalent pull was found in the workspace cache.")
    save_stage(run_id, "cache_lookup", "completed", completed_at=datetime.now(timezone.utc), duration_seconds=0, cache_hit=True)
    update_run(run_id, status="completed", result=result, activity_log=log)


def execute_pull_job(run_id: int, request_id: str, prompt: str, scope: dict, log: list[dict], workspace_id: int | None, cache_key: str | None, mode: str):
    log = PersistedActivityLog(run_id, log)
    started = time.monotonic()
    save_stage(run_id, "openai_web_search", "running")
    try:
        legacy.activity(log, "Pull stage started", f"Running {mode} OpenAI web-search and source-verification stage.")
        result = legacy.call_openai(prompt, scope, log)
        legacy.activity(log, "Backend request completed", "Returning live comps result to browser.")
        save_stage(
            run_id,
            "openai_web_search",
            "completed",
            completed_at=datetime.now(timezone.utc),
            duration_seconds=round(time.monotonic() - started, 3),
        )
        if workspace_id and cache_key:
            save_cache_entry(workspace_id, cache_key, PROMPT_VERSION, OPENAI_MODEL, "verified_result", result)
            legacy.activity(log, "Cache updated", "Saved verified pull result for identical reruns.", cache_key=cache_key[:12])
        update_run(
            run_id,
            status="completed",
            result=result,
            activity_log=result.get("activity_log", log),
        )
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        legacy.activity(log, "OpenAI request failed", "OpenAI returned an HTTP error.", status=exc.code)
        save_stage(run_id, "openai_web_search", "failed", completed_at=datetime.now(timezone.utc), duration_seconds=round(time.monotonic() - started, 3), error=body[:1000])
        update_run(run_id, "failed", error=body, activity_log=log)
    except urllib.error.URLError as exc:
        message = "The backend could not reach OpenAI after retries."
        legacy.activity(log, "OpenAI network request failed", message, error=str(exc))
        save_stage(run_id, "openai_web_search", "failed", completed_at=datetime.now(timezone.utc), duration_seconds=round(time.monotonic() - started, 3), error=str(exc))
        update_run(run_id, "failed", error=message, activity_log=log)
    except (TimeoutError, socket.timeout):
        message = "The live web-search request timed out before results came back."
        legacy.activity(log, "OpenAI request timed out", "The Responses API call exceeded the backend timeout.")
        save_stage(run_id, "openai_web_search", "failed", completed_at=datetime.now(timezone.utc), duration_seconds=round(time.monotonic() - started, 3), error=message)
        update_run(run_id, "failed", error=message, activity_log=log)
    except Exception as exc:
        legacy.activity(log, "Backend request failed", "Unhandled backend error.", error=str(exc))
        save_stage(run_id, "openai_web_search", "failed", completed_at=datetime.now(timezone.utc), duration_seconds=round(time.monotonic() - started, 3), error=str(exc))
        update_run(run_id, "failed", error=str(exc), activity_log=log)


def start_pull_job(prompt: str, scope: dict, files: list[UploadFile] | None, user, workspace_id: int, mode="standard", use_cache=True):
    workspace = require_workspace(workspace_id, user)
    mode = normalize_mode(mode)
    request_id, prepared_prompt, prepared_scope, log = prepare_scope(prompt, scope, files)
    prepared_scope["pullMode"] = mode
    cache_key = stable_cache_key(prepared_prompt, prepared_scope, mode)
    row = save_run(
        request_id=request_id,
        status="running",
        prompt=prepared_prompt,
        scope=prepared_scope,
        activity_log=log,
        workspace_id=workspace["id"],
        user_id=user["id"],
        mode=mode,
        cache_key=cache_key,
    )
    run_id = row["id"]
    save_stage(run_id, "cache_lookup", "running")
    if use_cache:
        cached = get_cache_entry(workspace["id"], cache_key, PROMPT_VERSION, OPENAI_MODEL, "verified_result")
        if cached:
            complete_from_cache(run_id, cached["payload_json"], log)
            return {"ok": True, "run_id": run_id, "request_id": request_id, "status": "completed", "activity_log": log, "cache_hit": True}
    legacy.activity(log, "Cache miss", "No completed equivalent pull found; starting live OpenAI workflow.", cache_key=cache_key[:12])
    save_stage(run_id, "cache_lookup", "completed", completed_at=datetime.now(timezone.utc), duration_seconds=0, cache_hit=False)
    update_run(run_id, "running", activity_log=log)
    thread = threading.Thread(
        target=execute_pull_job,
        args=(run_id, request_id, prepared_prompt, prepared_scope, log, workspace["id"], cache_key, mode),
        daemon=True,
    )
    thread.start()
    return {"ok": True, "run_id": run_id, "request_id": request_id, "status": "running", "activity_log": log, "cache_hit": False}


def run_pull_sync(prompt: str, scope: dict, files: list[UploadFile] | None, user, workspace_id: int, mode="standard"):
    response = start_pull_job(prompt, scope, files, user, workspace_id, mode=mode, use_cache=False)
    run_id = response["run_id"]
    while True:
        run = get_run(run_id, workspace_id)
        if run["status"] != "running":
            break
        time.sleep(0.5)
    if run["status"] == "completed":
        return {"ok": True, "run": run["result"]}
    return JSONResponse(
        status_code=502,
        content={"ok": False, "error": run["error"] or "Pull failed.", "activity_log": run["activity_log"] or []},
    )


def public_refinement(refinement):
    if not refinement:
        return None
    return {
        "id": refinement["id"],
        "run_id": refinement["run_id"],
        "instruction": refinement["instruction"],
        "model": refinement["model"],
        "openai_response_id": refinement["openai_response_id"],
        "created_at": refinement["created_at"],
    }


def latest_result_for_run(run):
    if not run or not run["workspace_id"]:
        return run["result"] if run else None, None
    refinement = get_latest_run_refinement(run["workspace_id"], run["id"])
    if refinement:
        return refinement["result"], refinement
    return run["result"], None


def merge_uploaded_dataset_metadata(current_result: dict | None, datasets: list[dict]):
    seen = set()
    merged = []
    for item in (current_result or {}).get("uploaded_datasets", []) or []:
        key = (item.get("name"), item.get("type"), item.get("row_count"))
        if key in seen:
            continue
        seen.add(key)
        merged.append(item)
    for item in datasets:
        metadata = {"name": item["name"], "type": item["type"], "row_count": item["row_count"]}
        key = (metadata["name"], metadata["type"], metadata["row_count"])
        if key in seen:
            continue
        seen.add(key)
        merged.append(metadata)
    return merged


def build_augmentation_instruction(datasets: list[dict], errors: list[dict], user_instruction: str):
    dataset_rows = [
        {
            "dataset": item["name"],
            "type": item["type"],
            "row_count": item["row_count"],
            "rows": item["rows"],
        }
        for item in datasets
    ]
    base_instruction = user_instruction.strip() or "No additional user instruction."
    return (
        "Augment the saved oncology comps result with the uploaded proprietary database exports.\n\n"
        "Workflow:\n"
        "1. Treat uploaded rows from Cortellis, GlobalData, or internal exports as candidate leads, not automatically verified comps.\n"
        "2. Search and verify relevant uploaded candidates against public primary sources where possible.\n"
        "3. Add missing relevant comps to the comps array, merge duplicate rows into existing comps, and fill missing fields when evidence supports it.\n"
        "4. Keep existing comps unless a row is a clear duplicate, out of scope, or structurally excluded by the original rules.\n"
        "5. Put rejected uploaded candidates into stripped_deals with a concise reason.\n"
        "6. Add or preserve custom_columns for database augmentation metadata, including database_source_file and database_augmentation_status when useful.\n"
        "7. Include a refinement_summary that explains how many uploaded files were reviewed, what was added, what was merged, and what was stripped.\n\n"
        f"User instruction:\n{base_instruction}\n\n"
        f"Uploaded dataset parse errors JSON:\n{json.dumps(errors, indent=2)}\n\n"
        f"Uploaded dataset rows JSON:\n{json.dumps(dataset_rows, indent=2)}\n\n"
        "Return the complete revised result JSON."
    )


def run_status_response(run_id: int, workspace_id: int | None = None):
    run = get_run(run_id, workspace_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found.")
    result, refinement = latest_result_for_run(run)
    stages = list_stages(run_id)
    base = {
        "status": run["status"],
        "run_id": run_id,
        "activity_log": run["activity_log"] or [],
        "stages": stages,
        "created_at": run["created_at"],
        "started_at": run["started_at"],
        "completed_at": run["completed_at"],
        "mode": run["mode"],
        "scope": run["scope"],
        "prompt": run["prompt"],
        "refinement": public_refinement(refinement),
    }
    if run["status"] == "completed":
        return {"ok": True, **base, "run": result}
    if run["status"] == "failed":
        return {"ok": False, **base, "error": run["error"] or "Pull failed."}
    return {"ok": True, **base}


@app.get("/api/runs")
def runs(limit: int = 25, user=Depends(get_current_user)):
    workspace = first_workspace_for_user(user)
    return {"ok": True, "runs": list_runs(limit=max(1, min(limit, 100)), workspace_id=workspace["id"])}


@app.get("/api/workspaces/{workspace_id}/runs")
def workspace_runs(workspace_id: int, limit: int = 25, user=Depends(get_current_user)):
    require_workspace(workspace_id, user)
    return {"ok": True, "runs": list_runs(limit=max(1, min(limit, 100)), workspace_id=workspace_id)}


@app.post("/api/pull")
def pull_json(request: PullRequest, user=Depends(get_current_user)):
    workspace = first_workspace_for_user(user)
    return start_pull_job(request.prompt, dict(request.scope), None, user, workspace["id"], request.mode, request.use_cache)


@app.post("/api/workspaces/{workspace_id}/pulls")
def workspace_pull_json(workspace_id: int, request: PullRequest, user=Depends(get_current_user)):
    return start_pull_job(request.prompt, dict(request.scope), None, user, workspace_id, request.mode, request.use_cache)


@app.post("/api/pull/form")
def pull_form(
    prompt: str = Form(""),
    scope: str = Form("{}"),
    mode: str = Form("standard"),
    use_cache: bool = Form(True),
    datasets: list[UploadFile] | None = File(None),
    user=Depends(get_current_user),
):
    try:
        parsed_scope = json.loads(scope)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid scope JSON: {exc}") from exc
    workspace = first_workspace_for_user(user)
    return start_pull_job(prompt, parsed_scope, datasets or [], user, workspace["id"], mode, use_cache)


@app.post("/api/workspaces/{workspace_id}/pulls/form")
def workspace_pull_form(
    workspace_id: int,
    prompt: str = Form(""),
    scope: str = Form("{}"),
    mode: str = Form("standard"),
    use_cache: bool = Form(True),
    datasets: list[UploadFile] | None = File(None),
    user=Depends(get_current_user),
):
    try:
        parsed_scope = json.loads(scope)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid scope JSON: {exc}") from exc
    return start_pull_job(prompt, parsed_scope, datasets or [], user, workspace_id, mode, use_cache)


@app.get("/api/runs/{run_id}")
def run_status(run_id: int, user=Depends(get_current_user)):
    run = get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found.")
    if run["workspace_id"]:
        require_workspace(run["workspace_id"], user)
    return run_status_response(run_id)


@app.get("/api/workspaces/{workspace_id}/runs/{run_id}")
def workspace_run_status(workspace_id: int, run_id: int, user=Depends(get_current_user)):
    require_workspace(workspace_id, user)
    return run_status_response(run_id, workspace_id)


@app.post("/api/workspaces/{workspace_id}/runs/{run_id}/refinements")
def refine_run(workspace_id: int, run_id: int, request: RefinementRequest, user=Depends(get_current_user)):
    require_workspace(workspace_id, user)
    run = get_run(run_id, workspace_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found.")
    if run["status"] != "completed" or not run["result"]:
        raise HTTPException(status_code=409, detail="Refinements are available after a run completes.")

    instruction = request.instruction.strip()
    current_result, _ = latest_result_for_run(run)
    log = []
    request_id = uuid.uuid4().hex[:10]
    legacy.activity(log, "Refinement request started", "Handling post-run refinement request.", request_id=request_id)
    try:
        refined_result = legacy.call_openai_refinement(current_result, instruction, run["scope"], log)
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        legacy.activity(log, "OpenAI refinement failed", "OpenAI returned an HTTP error.", status=exc.code)
        raise HTTPException(status_code=502, detail={"error": body, "activity_log": log}) from exc
    except urllib.error.URLError as exc:
        legacy.activity(log, "OpenAI refinement network request failed", "The backend could not reach OpenAI after retries.", error=str(exc))
        raise HTTPException(status_code=502, detail={"error": "The backend could not reach OpenAI after retries.", "activity_log": log}) from exc
    except (TimeoutError, socket.timeout) as exc:
        legacy.activity(log, "OpenAI refinement timed out", "The refinement request exceeded the backend timeout.")
        raise HTTPException(status_code=504, detail={"error": "The refinement request timed out before results came back.", "activity_log": log}) from exc
    except Exception as exc:
        legacy.activity(log, "Refinement request failed", "Unhandled backend error.", error=str(exc))
        raise HTTPException(status_code=500, detail={"error": str(exc), "activity_log": log}) from exc

    if "uploaded_datasets" not in refined_result and current_result:
        refined_result["uploaded_datasets"] = current_result.get("uploaded_datasets", [])
    refinement = save_run_refinement(
        workspace_id,
        run_id,
        user["id"],
        instruction,
        refined_result,
        refined_result.get("model"),
        refined_result.get("openai_response_id"),
    )
    return {
        "ok": True,
        "run_id": run_id,
        "run": refined_result,
        "scope": run["scope"],
        "prompt": run["prompt"],
        "mode": run["mode"],
        "created_at": run["created_at"],
        "completed_at": run["completed_at"],
        "refinement": public_refinement(refinement),
        "activity_log": log,
    }


@app.post("/api/workspaces/{workspace_id}/runs/{run_id}/augment/form")
def augment_run_with_uploads(
    workspace_id: int,
    run_id: int,
    instruction: str = Form(""),
    datasets: list[UploadFile] | None = File(None),
    user=Depends(get_current_user),
):
    require_workspace(workspace_id, user)
    run = get_run(run_id, workspace_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found.")
    if run["status"] != "completed" or not run["result"]:
        raise HTTPException(status_code=409, detail="Database augmentation is available after a run completes.")
    if not datasets:
        raise HTTPException(status_code=400, detail="Upload at least one .csv or .xlsx database export.")

    log = []
    request_id = uuid.uuid4().hex[:10]
    legacy.activity(log, "Database augmentation started", "Parsing uploaded database exports for post-run merge.", request_id=request_id)
    parsed_datasets, parse_errors = legacy.parse_uploaded_datasets([upload_to_legacy_field(file) for file in datasets], log)
    if not parsed_datasets:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "No supported rows were parsed from the uploaded files.",
                "dataset_errors": parse_errors,
                "activity_log": log,
            },
        )

    current_result, _ = latest_result_for_run(run)
    augmentation_scope = dict(run["scope"] or {})
    augmentation_scope["augmentationUploadedDatasets"] = [
        {"name": item["name"], "type": item["type"], "row_count": item["row_count"]}
        for item in parsed_datasets
    ]
    augmentation_scope["augmentationUploadedDatasetErrors"] = parse_errors
    model_instruction = build_augmentation_instruction(parsed_datasets, parse_errors, instruction)
    saved_instruction = "Augment with uploaded database exports: " + ", ".join(item["name"] for item in parsed_datasets)
    if instruction.strip():
        saved_instruction = f"{saved_instruction}. {instruction.strip()}"

    try:
        refined_result = legacy.call_openai_refinement(current_result, model_instruction, augmentation_scope, log)
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        legacy.activity(log, "Database augmentation failed", "OpenAI returned an HTTP error.", status=exc.code)
        raise HTTPException(status_code=502, detail={"error": body, "activity_log": log}) from exc
    except urllib.error.URLError as exc:
        legacy.activity(log, "Database augmentation network request failed", "The backend could not reach OpenAI after retries.", error=str(exc))
        raise HTTPException(status_code=502, detail={"error": "The backend could not reach OpenAI after retries.", "activity_log": log}) from exc
    except (TimeoutError, socket.timeout) as exc:
        legacy.activity(log, "Database augmentation timed out", "The augmentation request exceeded the backend timeout.")
        raise HTTPException(status_code=504, detail={"error": "The augmentation request timed out before results came back.", "activity_log": log}) from exc
    except Exception as exc:
        legacy.activity(log, "Database augmentation failed", "Unhandled backend error.", error=str(exc))
        raise HTTPException(status_code=500, detail={"error": str(exc), "activity_log": log}) from exc

    refined_result["uploaded_datasets"] = merge_uploaded_dataset_metadata(current_result, parsed_datasets)
    refined_result["activity_log"] = refined_result.get("activity_log", []) or log
    refinement = save_run_refinement(
        workspace_id,
        run_id,
        user["id"],
        saved_instruction[:4000],
        refined_result,
        refined_result.get("model"),
        refined_result.get("openai_response_id"),
    )
    return {
        "ok": True,
        "run_id": run_id,
        "run": refined_result,
        "scope": run["scope"],
        "prompt": run["prompt"],
        "mode": run["mode"],
        "created_at": run["created_at"],
        "completed_at": run["completed_at"],
        "refinement": public_refinement(refinement),
        "activity_log": log,
    }


@app.get("/api/workspaces/{workspace_id}/runs/{run_id}/stages")
def workspace_run_stages(workspace_id: int, run_id: int, user=Depends(get_current_user)):
    require_workspace(workspace_id, user)
    if not get_run(run_id, workspace_id):
        raise HTTPException(status_code=404, detail="Run not found.")
    return {"ok": True, "stages": list_stages(run_id)}


@app.post("/api/pull/sync")
def pull_json_sync(request: PullRequest, user=Depends(get_current_user)):
    workspace = first_workspace_for_user(user)
    return run_pull_sync(request.prompt, dict(request.scope), None, user, workspace["id"], request.mode)


def sse_line(event_id, payload):
    return f"id: {event_id}\nevent: activity\ndata: {json.dumps(payload, default=str)}\n\n"


@app.get("/api/workspaces/{workspace_id}/runs/{run_id}/events")
def run_events(workspace_id: int, run_id: int, user=Depends(get_current_user)):
    require_workspace(workspace_id, user)
    if not get_run(run_id, workspace_id):
        raise HTTPException(status_code=404, detail="Run not found.")

    def generate():
        sent = 0
        while True:
            run = get_run(run_id, workspace_id)
            if not run:
                break
            activity = run["activity_log"] or []
            for index, item in enumerate(activity[sent:], start=sent + 1):
                yield sse_line(index, item)
            sent = len(activity)
            if run["status"] in {"completed", "failed"}:
                yield sse_line("status", {"event": f"Run {run['status']}", "detail": "Run reached a terminal state.", "status": run["status"]})
                break
            time.sleep(2)

    return StreamingResponse(generate(), media_type="text/event-stream")


def add_sheet(workbook, title, rows):
    sheet = workbook.create_sheet(title=title[:31])
    header_fill = PatternFill("solid", fgColor="D9E8E5")
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if row_index == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    sheet.freeze_panes = "A2"
    widths = {}
    for row in rows:
        for index, value in enumerate(row, start=1):
            widths[index] = min(max(widths.get(index, 10), len(str(value or "")) + 2), 60)
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width


def first_value(mapping, *keys, default=None):
    for key in keys:
        value = mapping.get(key)
        if value is not None and value != "":
            return value
    return default


def custom_column_defs(result):
    columns = []
    for index, item in enumerate(result.get("custom_columns") or []):
        if isinstance(item, str):
            key = item.strip().lower().replace(" ", "_")
            label = item.strip()
        else:
            key = str(item.get("key") or "").strip()
            label = str(item.get("label") or key).strip()
        if not key:
            key = f"custom_{index + 1}"
        if not label:
            label = key.replace("_", " ").title()
        columns.append({"key": key, "label": label})
    return columns


def custom_value(deal, key):
    custom = deal.get("custom") or {}
    if isinstance(custom, dict):
        value = custom.get(key)
        if value is not None and value != "":
            return value
    return first_value(deal, key)


def comp_row(deal, custom_columns=None):
    row = [
        first_value(deal, "flag"),
        first_value(deal, "deal"),
        first_value(deal, "asset"),
        first_value(deal, "target_moa", "target", "target_company"),
        first_value(deal, "indication"),
        first_value(deal, "modality"),
        first_value(deal, "stage"),
        first_value(deal, "geography"),
        first_value(deal, "deal_type", "type"),
        first_value(deal, "announcement_date", "date"),
        first_value(deal, "announced_or_completed", "status", default="Announced"),
        first_value(deal, "upfront_usd_m", "upfront"),
        first_value(deal, "total_value_usd_m", "biobucks"),
        first_value(deal, "royalty"),
        first_value(deal, "primary_source_name", "source"),
        first_value(deal, "primary_source_url", "sourceUrl"),
        first_value(deal, "confidence"),
        first_value(deal, "analyst_note", "note"),
    ]
    for column in custom_columns or []:
        row.append(custom_value(deal, column["key"]))
    return row


def stripped_row(deal):
    return [
        first_value(deal, "deal"),
        first_value(deal, "reason"),
        first_value(deal, "source"),
        first_value(deal, "analyst_note", "note"),
    ]


def source_verification_row(deal):
    return [
        first_value(deal, "deal"),
        first_value(deal, "flag"),
        first_value(deal, "primary_source_name", "source"),
        first_value(deal, "primary_source_url", "sourceUrl"),
        first_value(deal, "analyst_note", "note"),
    ]


def workbook_for_run(run):
    result = run["result"] or {}
    comps = result.get("comps") or []
    stripped_deals = result.get("stripped_deals") or result.get("stripped") or []
    custom_columns = custom_column_defs(result)
    comp_headers = ["Flag", "Deal", "Asset", "Target", "Indication", "Modality", "Stage", "Geography", "Deal Type", "Announcement Date", "Status", "Upfront USD M", "Biobucks USD M", "Royalty", "Primary Source", "Source URL", "Confidence", "Analyst Note"]
    comp_headers.extend(column["label"] for column in custom_columns)
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "Summary", [["Field", "Value"], ["Run ID", run["id"]], ["Status", run["status"]], ["Mode", run["mode"]], ["Created", str(run["created_at"])], ["Completed", str(run["completed_at"] or "")], ["Summary", result.get("summary", "")]])
    add_sheet(workbook, "Comps", [comp_headers, *[
        comp_row(deal, custom_columns)
        for deal in comps
    ]])
    add_sheet(workbook, "Stripped Deals Audit", [["Deal", "Reason stripped", "Source", "Analyst note"], *[
        stripped_row(deal)
        for deal in stripped_deals
    ]])
    add_sheet(workbook, "Source Verification", [["Deal", "Flag", "Primary Source", "Source URL", "Analyst Note"], *[
        source_verification_row(deal)
        for deal in comps
    ]])
    add_sheet(workbook, "Uploaded Datasets", [["File", "Type", "Rows parsed"], *[
        [item.get("name"), item.get("type"), item.get("row_count")]
        for item in result.get("uploaded_datasets", [])
    ]])
    add_sheet(workbook, "Activity Log", [["Time", "Event", "Detail", "Metrics"], *[
        [item.get("time"), item.get("event"), item.get("detail"), " | ".join(f"{key}: {value}" for key, value in (item.get("metrics") or {}).items())]
        for item in (result.get("activity_log") or run["activity_log"] or [])
    ]])
    add_sheet(workbook, "Methodology", [["Methodology"], *[[item] for item in result.get("methodology", [])], ["Refinement summary"], [result.get("refinement_summary", "Not applicable")], ["Fallback / expansion warning"], [result.get("fallback_note", "Not applicable")]])
    add_sheet(workbook, "Prompt", [["Prompt"], [run["prompt"] or ""]])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@app.get("/api/workspaces/{workspace_id}/runs/{run_id}/exports/workbook")
@app.post("/api/workspaces/{workspace_id}/runs/{run_id}/exports/workbook")
def export_workbook(workspace_id: int, run_id: int, user=Depends(get_current_user)):
    require_workspace(workspace_id, user)
    run = get_run(run_id, workspace_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found.")
    if run["status"] != "completed" or not run["result"]:
        raise HTTPException(status_code=409, detail="Workbook export is available after a run completes.")
    result, _ = latest_result_for_run(run)
    export_run = {**run, "result": result}
    file_obj = workbook_for_run(export_run)
    filename = f"run-{run_id}-oncology-comps.xlsx"
    save_export_metadata(workspace_id, run_id, "workbook", "completed", filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", len(file_obj.getbuffer()))
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(file_obj, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
